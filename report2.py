from logging import debug
import pandas as pd
import numpy as np
from pandas.core.frame import DataFrame
import streamlit as st
import plotly.figure_factory as ff
import plotly.graph_objects as go
from plotly.io import to_image
import plotly.io as pio
import plotly.graph_objs.layout as go_layout
import openpyxl
from PIL import Image
import heapq #リストの最大値・最小値から順にn個の要素を取得
from io import BytesIO
import matplotlib.pyplot as plt
import japanize_matplotlib
from matplotlib.font_manager import FontProperties
from matplotlib import rcParams
import math




st.set_page_config(page_title='売り上げ分析（得意先別）レポート')
st.markdown('#### 売り上げ分析（得意先別）レポート自動作成')

#********************************データ準備

# ***今期出荷
@st.cache_data
def make_data_shukka(file):
    df_shukka = pd.read_excel(
    file, sheet_name='受注委託移動在庫生産照会', usecols=[3, 6, 15, 16, 45]) #index　ナンバー不要　index_col=0

    # *** 出荷月、受注月列の追加***
    df_shukka['出荷月'] = df_shukka['出荷日'].dt.month
    # ***INT型への変更***
    df_shukka[['金額', '出荷月', '受注月']] = df_shukka[[\
        '金額', '出荷月', '受注月']].fillna(0).astype('int64')

    return df_shukka

#今期目標
@st.cache_data
def make_data_target(file):
    df_target = pd.read_excel(file, sheet_name='Sheet1')
    df_target2 = df_target.set_index('得意先名')
    df_target2 = df_target2.T

    return df_target2

#***今期受注
@st.cache_data
def make_data_now(file):
    df_now = pd.read_excel(
    file, sheet_name='受注委託移動在庫生産照会', \
        usecols=[1, 3, 6, 8, 10, 14, 15, 16, 28, 31, 42, 50, 51, 52]) #index　ナンバー不要　index_col=0

    # *** 出荷月、受注月列の追加***
    df_now['出荷月'] = df_now['出荷日'].dt.month
    df_now['受注月'] = df_now['受注日'].dt.month
    df_now['商品コード2'] = df_now['商　品　名'].map(lambda x: x.split()[0]) #品番
    df_now['商品コード3'] = df_now['商　品　名'].map(lambda x: str(x)[0:2]) #頭品番
    df_now['張地'] = df_now['商　品　名'].map(lambda x: x.split()[2] if len(x.split()) >= 4 else '')

    # ***INT型への変更***
    df_now[['数量', '単価', '金額', '出荷倉庫', '原価金額', '出荷月', '受注月']] = \
        df_now[['数量', '単価', '金額', '出荷倉庫', '原価金額', '出荷月', '受注月']].fillna(0).astype('int64')
    #fillna　０で空欄を埋める

    return df_now

#***前期受注
@st.cache_data
def make_data_last(file):
    df_last = pd.read_excel(
    file, sheet_name='受注委託移動在庫生産照会', \
        usecols=[1, 3, 6, 8, 10, 14, 15, 16, 28, 31, 42, 50, 51, 52])
    df_last['出荷月'] = df_last['出荷日'].dt.month
    df_last['受注月'] = df_last['受注日'].dt.month
    df_last['商品コード2'] = df_last['商　品　名'].map(lambda x: x.split()[0])
    df_last['商品コード3'] = df_last['商　品　名'].map(lambda x: str(x)[0:2]) #頭品番
    df_last['張地'] = df_last['商　品　名'].map(lambda x: x.split()[2] if len(x.split()) >= 4 else '')

    df_last[['数量', '単価', '金額', '出荷倉庫', '原価金額', '出荷月', '受注月']] = \
        df_last[['数量', '単価', '金額', '出荷倉庫', '原価金額', '出荷月', '受注月']].fillna(0).astype('int64')
    #fillna　０で空欄を埋める

    return df_last

col1, col2, col3 = st.columns([3, 3, 3])

with col1:
    # ***ファイルアップロード 今期出荷***
    uploaded_file_shukka = st.file_uploader('出荷/今期', type='xlsx', key='snow')
    df_shukka = DataFrame()
    if uploaded_file_shukka:
        df_shukka = make_data_now(uploaded_file_shukka)

    else:
        st.info('出荷ベース（今期）')

    # ***ファイルアップロード　今期目標***
    uploaded_file_target = st.file_uploader('目標/今期', type='xlsx', key='target')
    df_target = DataFrame()
    if uploaded_file_target:
        df_target = make_data_target(uploaded_file_target)

    else:
        st.info('目標(今期)')

with col2:
    # ***ファイルアップロード 今期受注***
    uploaded_file_now = st.file_uploader('受注/今期', type='xlsx', key='now')
    df_now = DataFrame()
    if uploaded_file_now:
        df_now = make_data_now(uploaded_file_now)

    else:
        st.info('受注ベース（今期）')

    # ***ファイルアップロード　前期受注***
    uploaded_file_last = st.file_uploader('受注/前期', type='xlsx', key='last')
    df_last = DataFrame()
    if uploaded_file_last:
        df_last = make_data_last(uploaded_file_last)

    else:
        st.info('受注ベース（前期）')
        st.stop()

#******************************変数設定

colors10 = ['#DC143C', '#4682B4', '#66CDAA', '#FFD700', '#F0E68C',
            '#32CD32', '#F08080', '#FFA500', '#696969', '#FFFFF0']
            #Crimson/SteelBlue/MediumAquamarine/Gold/Khaki
            #LimeGreen/LightCoral/Orange/DimGray/Ivory
color_now = '#4682B4' #	SteelBlue
color_last = '#AFEEEE' # PaleTurquoise

colors2 = ['#4682B4', '#AFEEEE']
colors3 = ['#DC143C', '#4682B4', '#66CDAA'] # Crimson/SteelBlue/MediumAquamarine

#貼り付け先excel
target_file = 'report2_out.xlsx'

#*****************************得意先選定
with col3:
    cust_list = df_target.columns
    with st.form('得意先選択', clear_on_submit=False):

        # selectbox target ***
        cust_name = st.selectbox(
            '選択:',
            cust_list,
        )
        submitted = st.form_submit_button("決定")

if submitted:
    #*****************************表紙の作成
    wb = openpyxl.load_workbook('report2.xlsx')
    sh = wb['sheet1']

    #表紙得意先名
    sh['A1'] = cust_name

    #表紙期間
    start_day = df_now['受注日'].min()
    #時刻の削除
    start_day2 = str(start_day).split(' ')[0]
    start_day2 = start_day2.replace('-', '/')
    sh['E35'] = start_day2
    sh['H35'] = df_now['受注日'].max()

    wb.save(target_file)
    wb.close()

    #****************************データ準備
    #df本体の得意先での絞込み
    df_shukka_cust = df_shukka[df_shukka['得意先名']==cust_name]
    df_target_cust = df_target[cust_name]
    df_now_cust =df_now[df_now['得意先名']==cust_name]
    df_last_cust =df_last[df_last['得意先名']==cust_name]
    #合計金額
    cust_total_shukka = df_shukka_cust['金額'].sum()
    cust_total_target = df_target_cust.sum()
    cust_total_now = df_now_cust['金額'].sum()
    cust_total_last = df_last_cust['金額'].sum()

    progress_rate = f'{cust_total_shukka / cust_total_target * 100: 0.1f} %'
    diff_t = '{:,}'.format(cust_total_shukka - cust_total_target)

    total_comparison = f'{cust_total_now / cust_total_last * 100: 0.1f} %'
    diff = '{:,}'.format(cust_total_now - cust_total_last)

    #**************************************************************************関数
    #*********************************************************************line
    def make_line(df_now, df_target, label_now, label_target, title, cell):

        x = list(range(0, 12, 1))
        month_list=['10月', '11月', '12月', '1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月']
        #strにしないと順番が崩れる

        # figureオブジェクトを生成する
        fig = plt.figure(figsize=(6, 4))

        # axesオブジェクトをfigureオブジェクトに対して設定する
        ax = fig.add_subplot(1, 1, 1)

        # x軸を月名で設定する
        ax.set_xticks(x)
        ax.set_xticklabels(month_list)

        # axesオブジェクトに対して散布図のメソッドを設定する
        ax.plot(x, df_now, marker='.', ls='-',lw=1, label=label_now) #linestyle/linewidth
        ax.plot(x, df_target, marker='.', ls='--', lw=0.5, label=label_target)
        # ax.plot(x, df_now, marker='.', ls='-',lw=1, color='steelblue', label=label_now) #linestyle/linewidth
        # ax.plot(x, df_target, marker='.', ls='--', lw=0.5, color='lightblue', label=label_target)

        #グラフに値を書き込む
        for x, now, target in zip(x, df_now, df_target):
            val_now = f'{now/10000:.0f}'
            val_target = f'{target/10000:.0f}'
            ax.annotate(val_now, (x, now), textcoords='offset points', xytext=(0,10), ha='center', va='bottom', fontsize=6)
            ax.annotate(val_target, (x, target), textcoords='offset points', xytext=(0,-10), ha='center', va='top', fontsize=6)
            #ha 水平線上/va垂直線上　offsetpoints座標軸からの相対的な位置をポイント数で指定/xytextテキストの座標を(x,y)=(0,10)として指定
            #labelは凡例で使用

        #凡例
        ax.legend(loc='upper right', fontsize=6)

        #axの枠線の太さの設定
        ax.spines['top'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['right'].set_linewidth(0.3)

        #軸のフォントサイズ設定
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

        ax.set_title(title)

        plt.savefig('graph/fig.png')   # プロットしたグラフをファイルsin.pngに保存する

        #excelへの貼り付け
        #カレントディレクトリの画像指定
        img_dir = 'graph/fig.png'

        #画像オブジェクト作成
        img = Image.open(img_dir)

        #excelの読み込み
        wb = openpyxl.load_workbook(target_file)
        sh = wb['sheet1']

        #シートへの貼り付け
        img = openpyxl.drawing.image.Image(img_dir)
        sh.add_image(img, cell)

        wb.save(target_file)
        wb.close()

    #*******************************************************************bar_nowlast
    def make_bar_nowlast(now, last, title, cell):

        # figureオブジェクトを生成する
        fig = plt.figure(figsize=(6, 4))

        # axesオブジェクトをfigureオブジェクトに対して設定する
        ax = fig.add_subplot(1, 1, 1)

        x = ['今期', '前期']
        y = [now, last]
        # axesオブジェクトに対して棒グラフの設定
        len_x = np.arange(len(x))

        #barグラフの作成
        rect = ax.bar(x, y, color=colors2)
        ax.set_xticks(len_x)
        ax.set_xticklabels(x)

        # グラフに値を書き込む
        def autolabel(rects):
            for rect in rects:
                height = rect.get_height() #y軸の値
                ax.annotate(f'{height/10000: .0f}',
                        xy=(rect.get_x() + rect.get_width() / 2, height), #xy textの書き込み座標 x軸　barのwidth
                        xytext=(0, 3), #text表示位置 barのy軸方向に+3
                        textcoords="offset points", #座標が、データ座標系ではなく、テキスト座標系
                        ha='center', va='bottom', #textの水平位置、垂直位置
                        fontsize=6)
        autolabel(rect)

        #axの枠線の太さの設定
        ax.spines['top'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['right'].set_linewidth(0.3)

        #軸のフォントサイズ設定
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

        ax.set_title(title)

        plt.savefig('graph/fig.png')   # プロットしたグラフをファイルsin.pngに保存する

        #excelへの貼り付け
        #カレントディレクトリの画像指定
        img_dir = 'graph/fig.png'

        #画像オブジェクト作成
        img = Image.open(img_dir)

        wb = openpyxl.load_workbook(target_file)
        sh = wb['sheet1']

        #売上（累計）グラフ
        img = openpyxl.drawing.image.Image(img_dir)
        sh.add_image(img, cell)

        wb.save(target_file)
        wb.close()

    #*******************************************************************bar_nowlast halfsize
    def make_bar_nowlast_half(now, last, title, cell, ymax='none'):

        # figureオブジェクトを生成する
        fig = plt.figure(figsize=(3, 4))

        # axesオブジェクトをfigureオブジェクトに対して設定する
        ax = fig.add_subplot(1, 1, 1)

        x = ['今期', '前期']
        y = [now, last]
        # axesオブジェクトに対して棒グラフの設定
        len_x = np.arange(len(x))

        #barグラフの作成
        rect = ax.bar(x, y, color=colors2)
        ax.set_xticks(len_x)
        ax.set_xticklabels(x)

        # グラフに値を書き込む
        def autolabel(rects):
            for rect in rects:
                height = rect.get_height() #y軸の値
                ax.annotate(f'{height/10000: .0f}',
                        xy=(rect.get_x() + rect.get_width() / 2, height), #xy textの書き込み座標 x軸　barのwidth
                        xytext=(0, 3), #text表示位置 barのy軸方向に+3
                        textcoords="offset points", #座標が、データ座標系ではなく、テキスト座標系
                        ha='center', va='bottom', #textの水平位置、垂直位置
                        fontsize=6)
        autolabel(rect)

        #axの枠線の太さの設定
        ax.spines['top'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['right'].set_linewidth(0.3)

        #軸のフォントサイズ設定
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

        if ymax != 'none':
            plt.ylim(0, ymax)

        ax.set_title(title)

        plt.savefig('graph/fig.png')   # プロットしたグラフをファイルsin.pngに保存する

        #excelへの貼り付け
        #カレントディレクトリの画像指定
        img_dir = 'graph/fig.png'

        #画像オブジェクト作成
        img = Image.open(img_dir)

        wb = openpyxl.load_workbook(target_file)
        sh = wb['sheet1']

        #売上（累計）グラフ
        img = openpyxl.drawing.image.Image(img_dir)
        sh.add_image(img, cell)

        wb.save(target_file)
        wb.close()

    #**********************************************************************bar_malti
    def make_bar_multi(df, title, cell):
        # figureオブジェクトを生成する
        fig = plt.figure(figsize=(6, 4))

        # axesオブジェクトをfigureオブジェクトに対して設定する
        ax = fig.add_subplot(1, 1, 1)

        # x軸の目盛のラベルの位置を変数xで保持する。
        x = np.arange(len(df.index))

        #グラフの作成
        bar_now = ax.bar(x, df['今期'], align="edge", width=-0.3, label='今期',\
                        tick_label=df.index)
        bar_last = ax.bar(x, df['前期'], align="edge", width= 0.3, label='前期')

        #ラベルが重ならないように45度に設定
        plt.tick_params(axis='x', labelrotation=45)
        #グラフ下部に余白作成
        plt.subplots_adjust(bottom=0.2)



        # グラフに値を書き込む
        def autolabel(bar):
            for rect in bar:
                height = rect.get_height() #y軸の値
                
                #10000以上以下で処理を変える
                height2 = ''
                if height >= 10000:
                    height2 = f'{height/10000: .0f}'
                elif height < 1:
                    height2 = f'{height: .2f}'    
                else:
                    height2 = f'{height: .0f}'   
                

                ax.annotate(height2,
                        xy=(rect.get_x() + rect.get_width() / 2, height), #xy textの書き込み座標 x軸　barのwidth
                        xytext=(0, 3), #text表示位置 barのy軸方向に+3
                        textcoords="offset points", #座標が、データ座標系ではなく、テキスト座標系
                        ha='center', va='bottom', #textの水平位置、垂直位置
                        fontsize=6)
        autolabel(bar_now)
        autolabel(bar_last)

        # 凡例を表示する。
        ax.legend(loc="upper right", fontsize=8)

        #axの枠線の太さの設定
        ax.spines['top'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['right'].set_linewidth(0.3)

        #軸のフォントサイズ設定
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

        ax.set_title(title)

        plt.savefig('graph/fig.png')   # プロットしたグラフをファイルsin.pngに保存する

        #excelへの貼り付け
        #カレントディレクトリの画像指定
        img_dir = 'graph/fig.png'

        #画像オブジェクト作成
        img = Image.open(img_dir)

        wb = openpyxl.load_workbook(target_file)
        sh = wb['sheet1']

        #売上（累計）グラフ
        img = openpyxl.drawing.image.Image(img_dir)
        sh.add_image(img, cell)

        wb.save(target_file)
        wb.close()    

    #**********************************************************************pie
    def make_pie(now, labels, title, cell):

        #マイナスの値を0に変換　マイナスがあるとグラフ作成時にエラーが出る
        now = [0 if i < 0 else i for i in now]

        # figureオブジェクトを生成する
        fig = plt.figure(figsize=(3, 4))

        # axesオブジェクトをfigureオブジェクトに対して設定する
        ax = fig.add_subplot(1, 1, 1)

        ax.pie(now, labels=labels, autopct="%1.1f %%", textprops={"size": 7}, startangle=90,
               pctdistance=0.8, labeldistance=1.1) #開始角度/％表示位置中心から/ラベル表示位置

        #axの枠線の太さの設定
        ax.spines['top'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['right'].set_linewidth(0.3)

        #軸のフォントサイズ設定
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

        ax.set_title(title)

        plt.savefig('graph/fig.png')   # プロットしたグラフをファイルsin.pngに保存する

        #excelへの貼り付け
        #カレントディレクトリの画像指定
        img_dir = 'graph/fig.png'

        #画像オブジェクト作成
        img = Image.open(img_dir)

        wb = openpyxl.load_workbook(target_file)
        sh = wb['sheet1']

        #売上（累計）グラフ
        img = openpyxl.drawing.image.Image(img_dir)
        sh.add_image(img, cell)

        wb.save(target_file)
        wb.close()

    #************************************************************************0出荷ベース

    month_list = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    columns_list = ['目標', '出荷/今期', '対目標差', '対目標比']

    shukka_list = []
    target_diff_list = []
    target_rate_list = []

    target_num = 0
    for month in month_list:

        shukka_sum = df_shukka_cust[df_shukka_cust['出荷月'].isin([month])]['金額'].sum()
        target = list(df_target_cust)[target_num]

        target_diff = shukka_sum - target
        target_rate = f'{shukka_sum / target * 100: 0.1f} %'

        shukka_list.append('{:,}'.format(shukka_sum))
        target_diff_list.append(target_diff)
        target_rate_list.append(target_rate)

        target_num += 1

    df_month = pd.DataFrame(list(zip(\
        list(df_target_cust), shukka_list, target_diff_list, target_rate_list, \
            )), columns=columns_list, index=month_list)

    #*****月別可視化
    df_month2 = df_month.copy()

    #グラフ用にint化
    df_month2['出荷/今期2'] = df_month2['出荷/今期'].apply(lambda x: int(x.replace(',', '')))

    #******グラフの追加
    make_line(df_month2['出荷/今期2'], df_month2['目標'], \
              '出荷/今期', '目標', '売上/目標 月別推移', 'A40')

    #************************************************出荷目標/累計可視化

    #グラフ用にint化
    df_month2['累計/目標'] = df_month2['目標'].cumsum()
    df_month2['累計/出荷/今期2'] = df_month2['出荷/今期2'].cumsum()

    #table用にdiffとrate追加
    df_month2['累計/目標差'] = df_month2['累計/出荷/今期2'] - df_month2['累計/目標']
    df_month2['累計/目標比'] = df_month2['累計/出荷/今期2'] / df_month2['累計/目標']
    df_month2['累計/目標比'] = df_month2['累計/目標比'].apply(lambda x: '{:.2f}'.format(x))

    #グラフ作成
    make_line(df_month2['累計/出荷/今期2'], df_month2['累計/目標'], \
              '累計/出荷/今期', '累計/目標', '売上/目標 累計月別推移', 'A57')

    #excelへの直接書き込み
    wb = openpyxl.load_workbook(target_file)
    sh = wb['sheet1']

    #************excel
    #売上（累計）書き込み
    sh['C73'] = progress_rate
    sh['H73'] = diff_t

    wb.save(target_file)
    wb.close()


# #     #************************************************************************1年間累計売上

    make_bar_nowlast(cust_total_now, cust_total_last, '売上/累計', 'A80')

    #excelへの直接書き込み
    wb = openpyxl.load_workbook(target_file)
    sh = wb['sheet1']

    #************excel
    #売上（累計）書き込み
    sh['C97'] = total_comparison
    sh['H97'] = diff

    wb.save(target_file)
    wb.close()

# # ***************************************************2売上月別推移:

    month_list = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    columns_list = ['今期', '前期', '対前年差', '対前年比']

    earnings_now = []
    earnings_last = []
    earnings_diff = []
    earnings_rate = []

    for month in month_list:
        earnings_month_now = df_now_cust[df_now_cust['受注月'].isin([month])]['金額'].sum()
        earnings_month_last = df_last_cust[df_last_cust['受注月'].isin([month])]['金額'].sum()
        earnings_diff_culc = earnings_month_now - earnings_month_last
        earnings_rate_culc = f'{earnings_month_now / earnings_month_last * 100: 0.1f} %'

        earnings_now.append('{:,}'.format(earnings_month_now))
        earnings_last.append('{:,}'.format(earnings_month_last))
        earnings_diff.append('{:,}'.format(earnings_diff_culc))
        earnings_rate.append(earnings_rate_culc)

    df_earnings_month = pd.DataFrame(list(zip(earnings_now, earnings_last, earnings_diff, earnings_rate)),
                                     columns=columns_list, index=month_list)

    #グラフ用にintのデータを用意
    df_earnings_month2 = df_earnings_month.copy()
    df_earnings_month2['今期'] = df_earnings_month2['今期'].apply(lambda x: int(x.replace(',', '')))
    df_earnings_month2['前期'] = df_earnings_month2['前期'].apply(lambda x: int(x.replace(',', '')))

    make_line(df_earnings_month2['今期'], df_earnings_month2['前期'], '今期', '前期',
              '月別売上', 'A101')

# **********************************************************3平均成約単価:
    month_list = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    columns_list = ['今期', '前期', '対前年差', '対前年比']

    order_num_now = []
    for num in df_now_cust['伝票番号']:
        num2 = num.split('-')[0]
        order_num_now.append(num2)
    df_now_cust['order_num'] = order_num_now

    order_num_last = []
    for num in df_last_cust['伝票番号']:
        num2 = num.split('-')[0]
        order_num_last.append(num2)
    df_last_cust['order_num'] = order_num_last


    earnings_now = []
    earnings_last = []
    earnings_diff = []
    earnings_rate = []

    for month in month_list:
        earnings_month_now = df_now_cust[df_now_cust['受注月'].isin([month])]
        order_sum_now = earnings_month_now.groupby('order_num')['金額'].sum()
        order_mean_now = order_sum_now.mean()

        earnings_month_last = df_last_cust[df_last_cust['受注月'].isin([month])]
        order_sum_last = earnings_month_last.groupby('order_num')['金額'].sum()
        order_mean_last = order_sum_last.mean()

        order_mean_diff = order_mean_now - order_mean_last
        if order_mean_last == 0:
            order_mean_rate = '0%'
        else:
            order_mean_rate = f'{(order_mean_now / order_mean_last)*100: 0.1f} %'

        earnings_now.append(order_mean_now)
        earnings_last.append(order_mean_last)
        earnings_diff.append(order_mean_diff)
        earnings_rate.append(order_mean_rate)

    df_mean_earninngs_month = \
        pd.DataFrame(list(zip(earnings_now, earnings_last, earnings_diff, earnings_rate)), \
                     columns=columns_list, index=month_list)

    df_mean_earninngs_month.fillna(0, inplace=True)
    df_mean_earninngs_month['今期'] = \
        df_mean_earninngs_month['今期'].map(lambda x: '{:,}'.format(int(x)))
    df_mean_earninngs_month['前期'] = \
        df_mean_earninngs_month['前期'].map(lambda x: '{:,}'.format(int(x)))
    df_mean_earninngs_month['対前年差'] = \
        df_mean_earninngs_month['対前年差'].map(lambda x: '{:,}'.format(int(x)))

    #グラフ用にintのデータを用意
    df_mean_earninngs_month2 = df_mean_earninngs_month.copy()
    df_mean_earninngs_month2['今期'] = \
        df_mean_earninngs_month2['今期'].apply(lambda x: int(x.replace(',', '')))
    df_mean_earninngs_month2['前期'] = \
        df_mean_earninngs_month2['前期'].apply(lambda x: int(x.replace(',', '')))

    make_line(df_mean_earninngs_month2['今期'], df_mean_earninngs_month2['前期'],
              '今期', '前期', '平均成約単価', 'A117')

    wb = openpyxl.load_workbook(target_file)
    sh = wb['sheet1']

    #数値入力
    #合計平均単価算出
    val_now = df_now_cust['金額'].sum() / int(df_now_cust['order_num'].nunique())
    val_last = df_last_cust['金額'].sum() / int(df_last_cust['order_num'].nunique())

    if math.isnan(val_now):
        val_now = '0'
    else:
        val_now = '{:,}'.format(int(val_now))

    if math.isnan(val_last):
        val_last = '0'
    else:
        val_last = '{:,}'.format(int(val_last))

    sh['C134'] = val_now
    sh['H134'] = val_last

    wb.save(target_file)
    wb.close()

# ***************************************************************4LD比率:
    #***********living
    living_now = df_now_cust[df_now_cust['商品分類名2'].isin(
        ['クッション', 'リビングチェア', 'リビングテーブル'])]['金額'].sum()
    living_last = df_last_cust[df_last_cust['商品分類名2'].isin(
        ['クッション', 'リビングチェア', 'リビングテーブル'])]['金額'].sum()

    l_diff = living_now-living_last
    l_ratio = f'{living_now/living_last*100:0.1f} %'

    #**********dining
    dining_now = df_now_cust[df_now_cust['商品分類名2'].isin(
        ['ダイニングテーブル', 'ダイニングチェア', 'ベンチ'])]['金額'].sum()
    dining_last = df_last_cust[df_last_cust['商品分類名2'].isin(
        ['ダイニングテーブル', 'ダイニングチェア', 'ベンチ'])]['金額'].sum()

    d_diff = dining_now-dining_last
    d_ratio = f'{dining_now/dining_last*100:0.1f} %'

    #グラフのy軸を揃えるための作業
    sum_list = [living_now, living_last, dining_now, dining_last]
    sum_max = heapq.nlargest(1, sum_list) #listの最大値から何個取得
    sum_max = round(sum_max[0] / 100000) * 100000 + 300000 #max値を１０万単位にして30万プラス

    #リビング可視化
    make_bar_nowlast_half(living_now, living_last, 'リビング売上', 'A159', sum_max)
    #ダイニング可視化
    make_bar_nowlast_half(dining_now, dining_last, 'ダイニング売上', 'F159', sum_max)

    #excelに直接書き込み
    wb = openpyxl.load_workbook(target_file)
    sh = wb['sheet1']

    #文字入力
    sh['C175'] = l_ratio
    sh['C176'] = l_diff
    sh['A177'] = 'クッション/リビングチェア/リビングテーブル'

        #文字入力
    sh['H175'] = d_ratio
    sh['H176'] = d_diff
    sh['F177'] = 'ダイニングテーブル/ダイニングチェア/ベンチ'

    wb.save(target_file)
    wb.close()


#*****************LD比率今期
    else_now = df_now_cust[df_now_cust['商品分類名2'].isin(
        ['キャビネット類', 'その他テーブル', '雑品・特注品', 'その他椅子','デスク', '小物・その他'])]['金額'].sum()
    else_last = df_last_cust[df_last_cust['商品分類名2'].isin(
        ['キャビネット類', 'その他テーブル', '雑品・特注品', 'その他椅子','デスク', '小物・その他'])]['金額'].sum()

    comp_now_list = [living_now, dining_now, else_now]
    comp_now_index = ['リビング', 'ダイニング', 'その他']
    comp_now_columns = ['分類']
    df_comp_now = pd.DataFrame(comp_now_index, columns=comp_now_columns)
    df_comp_now['金額'] = comp_now_list

    #*****************ld比率/前年
    comp_last_list = [living_last, dining_last, else_last]
    comp_last_index = ['リビング', 'ダイニング', 'その他']
    comp_last_columns = ['分類']
    df_comp_last = pd.DataFrame(comp_last_index, columns=comp_last_columns)
    df_comp_last['金額'] = comp_last_list

    #円グラフ　今期
    make_pie(comp_now_list, comp_now_index, 'LD比率/今期', 'A180')

    #円グラフ　前期
    make_pie(comp_last_list, comp_last_index, 'LD比率/前期', 'F180')


# # #************************************************************5LD別シリーズ別構成比　リビング

    df_now_cust_cate = df_now_cust[df_now_cust['商品分類名2'].isin(
        ['クッション', 'リビングチェア', 'リビングテーブル'])]
    df_last_cust_cate = df_last_cust[df_last_cust['商品分類名2'].isin(
        ['クッション', 'リビングチェア', 'リビングテーブル'])]

    index = []
    now_result = []
    last_result = []
    diff = []
    ratio = []
    series_list = df_now_cust_cate['シリーズ名'].unique()

    for series in series_list:
        index.append(series)
        now_culc = df_now_cust_cate[df_now_cust_cate['シリーズ名']==series]['金額'].sum()
        last_culc = df_last_cust_cate[df_last_cust_cate['シリーズ名']==series]['金額'].sum()
        now_result.append(now_culc)
        last_result.append(last_culc)
        diff_culc = '{:,}'.format(now_culc - last_culc)
        diff.append(diff_culc)
        ratio_culc = f'{now_culc/last_culc*100:0.1f} %'
        ratio.append(ratio_culc)
    df_result = pd.DataFrame(list(zip(now_result, last_result, ratio, diff)), index=index, columns=['今期', '前期', '対前年比', '対前年差'])

    #**************今期
    #シリーズ数１０で足切り
    df_result_now = df_result['今期'].sort_values(ascending=False)
    df_result_last = df_result['前期'].sort_values(ascending=False)

    #result2と3をconcat
    df_resultm = pd.concat([df_result_now, df_result_last], axis=1, join='outer')
    df_resultm = df_resultm.sort_values('今期', ascending=False)

    if len(df_resultm) > 10:
        df_resultm2 = df_resultm[0:10]

    if len(df_result_now) > 10:
        df_result_now = df_result_now[0:10]

    if len(df_result_last) > 10:
        df_result_last = df_result_last[0:10]    


    # #シリーズ数の算出
    # len_now = len(df_result_now)
    # len_last = len(df_result_last)

    # #色リスト（シリーズ数に合わせた）の作成
    # selected_colors_now = []
    # for color in colors10[:len_now]:
    #     selected_colors_now.append(color)
  

    # #*********前年比棒グラフ リビング********
    make_bar_multi(df_resultm, 'シリーズ別売上/リビング', 'A199')


    #*****************リビング円グラフ
    make_pie(df_result_now, df_result_now.index, 'シリーズ別構成比/リビング/今期', 'A216')
    make_pie(df_result_last, df_result_last.index, 'シリーズ別構成比/リビング/前期', 'F216')


#**************************************6LD別シリーズ別構成比　ダイニング

    df_now_cust_cate = df_now_cust[df_now_cust['商品分類名2'].isin(
        ['ダイニングテーブル', 'ダイニングチェア', 'ベンチ'])]
    df_last_cust_cate = df_last_cust[df_last_cust['商品分類名2'].isin(
        ['ダイニングテーブル', 'ダイニングチェア', 'ベンチ'])]

    index = []
    now_result = []
    last_result = []
    diff = []
    ratio = []
    series_list = df_now_cust_cate['シリーズ名'].unique()

    for series in series_list:
        index.append(series)
        now_culc = df_now_cust_cate[df_now_cust_cate['シリーズ名']==series]['金額'].sum()
        last_culc = df_last_cust_cate[df_last_cust_cate['シリーズ名']==series]['金額'].sum()
        now_result.append(now_culc)
        last_result.append(last_culc)
        diff_culc = '{:,}'.format(now_culc - last_culc)
        diff.append(diff_culc)
        ratio_culc = f'{now_culc/last_culc*100:0.1f} %'
        ratio.append(ratio_culc)
    df_result = pd.DataFrame(list(zip(now_result, last_result, ratio, diff)),
                             index=index, columns=['今期', '前期', '対前年比', '対前年差'])

    #シリーズ数１０で足切り
    df_result_now = df_result['今期'].sort_values(ascending=False)
    df_result_last = df_result['前期'].sort_values(ascending=False)

    #result2と3をconcat
    df_resultm = pd.concat([df_result_now, df_result_last], axis=1, join='outer')
    df_resultm.sort_values('今期', ascending=False)

    if len(df_resultm) > 10:
        df_resultm = df_resultm[0:10]

    if len(df_result_now) > 10:
        df_result_now = df_result_now[0:10]

    if len(df_result_last) > 10:
        df_result_last = df_result_last[0:10]        

    #シリーズ数の算出
    len_df = len(df_resultm)

    #色リスト（シリーズ数に合わせた）の作成
    selected_colors = []
    for color in colors10[:len_df]:
        selected_colors.append(color)

    # #*********前年比棒グラフ ダイニング**************
    make_bar_multi(df_resultm, 'シリーズ別売上/ダイニング', 'A238')

    #*****************リビング円グラフ
    make_pie(df_result_now, df_result_now.index, 'シリーズ別構成比/ダイニング/今期', 'A255')
    make_pie(df_result_last, df_result_last.index, 'シリーズ別構成比/ダイニング/前期', 'F255')


# #***********************************************************************7塗色

    #***************塗色別構成比推移
    df_now_cust = df_now_cust.dropna(subset=['塗色CD'])
    df_last_cust = df_last_cust.dropna(subset=['塗色CD'])

    # ***塗色別売り上げ ***
    color_now = df_now_cust.groupby('塗色CD')['金額'].sum().sort_values(ascending=False) #降順
    #color_now2 = color_now.apply('{:,}'.format) #数値カンマ区切り　注意strになる　グラフ作れなくなる

    #***塗色別売り上げ ***
    color_last = df_last_cust.groupby('塗色CD')['金額'].sum().sort_values(ascending=False) #降順
    #color_last2 = color_now.apply('{:,}'.format) #数値カンマ区切り　注意strになる　グラフ作れなくなる

    color_now.rename('今期', inplace=True)
    color_last.rename('前期', inplace=True)
    #今期と前期を一体化
    df_color = pd.concat([color_now, color_last], join='outer', axis=1)
    df_color = df_color.fillna(0)

    #構成比df作成
    df_ratio = df_color/df_color.sum()
    #text用に小数点以下２ケタのカラムを用意
    df_ratio['今期2'] = df_ratio['今期'].apply(lambda x: f'{x: .1f}')
    df_ratio['前期2'] = df_ratio['前期'].apply(lambda x: f'{x: .1f}')

    # #*********前年比棒グラフ**************
    make_bar_multi(df_ratio, '塗色別構成比推移', 'A277')

    #*****************円グラフ
    make_pie(color_now, color_now.index, '塗色別構成比/今期', 'A294')
    make_pie(color_last, color_last.index, '塗色別構成比/前期', 'F294')


#*****************************************************************8張地　リビングチェア
     # *** selectbox***
    category = 'リビングチェア'

    categorybase_now = df_now_cust[df_now_cust['商品分類名2']==category]
    categorybase_last = df_last_cust[df_last_cust['商品分類名2']==category]

    categorybase_cust_now = categorybase_now[categorybase_now['張地'] != ''] #空欄を抜いたdf作成
    categorybase_cust_last = categorybase_last[categorybase_last['張地'] != '']

    # ***張地別数量 ***
    fabric_now = categorybase_cust_now.groupby('張地')['数量'].sum().sort_values(ascending=False) #降順
    #fabric2 = fabric_now.apply('{:,}'.format) #数値カンマ区切り　注意strになる　グラフ作れなくなる
    # ***張地別売り上げ ***
    fabric_last = categorybase_cust_last.groupby('張地')['数量'].sum().sort_values(ascending=False) #降順
    #fabric2 = fabric_now.apply('{:,}'.format) #数値カンマ区切り　注意strになる　グラフ作れなくなる

    fabric_now.rename('今期', inplace=True)
    fabric_last.rename('前期', inplace=True)
    #今期と前期を一体化
    df_fabric = pd.concat([fabric_now, fabric_last], join='outer', axis=1)
    df_fabric = df_fabric.fillna(0)

     #布数を１０で足切り colorsの関係
    if len(df_fabric) >10:
        df_fabric = df_fabric[0:10]
    len_df = len(df_fabric)
    selected_colors = colors10[0:len_df]

#     #*******************張地推移
    # #*********前年比棒グラフ ダイニング**************
    make_bar_multi(df_fabric, '張地別数量推移/リビングチェア', 'A316')

    #*****************リビング円グラフ
    make_pie(fabric_now, fabric_now.index, '張地別構成比/今期', 'A333')
    make_pie(fabric_last, fabric_last.index, '張地別構成比/前期', 'F333')


    with col3:
        #download
        st.image('download.png', width=70)
        st.caption('ダウンロードの準備が完了しました')

        # Excelファイルを読み込み、バイナリデータに変換する
        wb = openpyxl.load_workbook(filename='report2_out.xlsx')
        stream = BytesIO()
        wb.save(stream)
        data = stream.getvalue()

        # ダウンロードボタンを表示する
        st.download_button(label='ダウンロード開始', data=data, file_name=f'report_{cust_name}.xlsx')

        #link
        link = '[home](https://cocosan1-hidastreamlit3-linkpage-j6o5p1.streamlit.app/)'
        st.markdown(link, unsafe_allow_html=True)
        st.caption('homeに戻る')

